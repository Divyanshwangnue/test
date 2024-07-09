from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, jsonify
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, SelectField, SubmitField
from wtforms.validators import DataRequired
import imaplib
import poplib
import email
from email.policy import default
from bs4 import BeautifulSoup
import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os
import shutil
import uuid

app = Flask(__name__)
app.secret_key = 'supersecretkey'

class EmailForm(FlaskForm):
    email = StringField('Email', validators=[DataRequired()])
    password = PasswordField('Password', validators=[DataRequired()])
    protocol = SelectField('Protocol', choices=[('imap', 'IMAP'), ('pop3', 'POP3')], validators=[DataRequired()])
    server = StringField('Server', validators=[DataRequired()])
    port = StringField('Port')
    submit = SubmitField('Extract FTP Details')

def fetch_emails_imap(username, password, server, port):
    try:
        if not port:
            port = 993
        mail = imaplib.IMAP4_SSL(server, int(port))
        mail.login(username, password)
        mail.select('inbox')
        status, data = mail.search(None, 'ALL')
        mail_ids = data[0].split()
        emails = []
        for i in tqdm(mail_ids, desc="Fetching emails via IMAP", unit="email"):
            status, data = mail.fetch(i, '(RFC822)')
            emails.append(data[0][1])
        mail.logout()
        return emails
    except Exception as e:
        raise RuntimeError(f"IMAP error: {e}")

def fetch_emails_pop3(username, password, server, port):
    try:
        if not port:
            port = 995
        mail = poplib.POP3_SSL(server, int(port))
        mail.user(username)
        mail.pass_(password)
        num_messages = len(mail.list()[1])
        emails = []
        for i in tqdm(range(num_messages), desc="Fetching emails via POP3", unit="email"):
            response, lines, octets = mail.retr(i + 1)
            msg_data = b'\r\n'.join(lines)
            emails.append(msg_data)
        mail.quit()
        return emails
    except Exception as e:
        raise RuntimeError(f"POP3 error: {e}")

def parse_ftp_details_from_html(html_content):
    soup = BeautifulSoup(html_content, 'html.parser')
    account_info_table = soup.find('table', {'class': 'MsoNormalTable'})
    if not account_info_table:
        return None
    
    rows = account_info_table.find_all('tr')
    ftp_details = {}
    
    for row in rows:
        cells = row.find_all('td')
        if len(cells) == 3:
            label = cells[0].get_text(strip=True)
            key = cells[1].get_text(strip=True)
            value = cells[2].get_text(strip=True)
            if key and value:
                ftp_details[key] = value
    
    return ftp_details if ftp_details else None

def extract_ftp_details_from_emails(emails):
    all_details = []
    for raw_email in tqdm(emails, desc="Extracting FTP details", unit="email"):
        msg = email.message_from_bytes(raw_email, policy=default)
        for part in msg.walk():
            if part.get_content_type() == 'text/html':
                html_content = part.get_payload(decode=True)
                try:
                    html_content = html_content.decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        html_content = html_content.decode('latin-1')
                    except UnicodeDecodeError:
                        continue

                details = parse_ftp_details_from_html(html_content)
                if details:
                    details['Sender'] = msg['From']
                    details['Date'] = msg['Date']
                    all_details.append(details)
    return all_details

def save_to_excel(details_list, output_file):
    df = pd.DataFrame(details_list)
    df.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active

    # Adding dropdown for date column
    date_col = get_column_letter(df.columns.get_loc('Date') + 1)
    dv = DataValidation(type="list", formula1=f'"{",".join(df["Date"].unique())}"', showDropDown=True)
    ws.add_data_validation(dv)
    dv.add(f'{date_col}2:{date_col}{len(df) + 1}')

    wb.save(output_file)

@app.route('/', methods=['GET', 'POST'])
def index():
    form = EmailForm()
    if form.validate_on_submit():
        username = form.email.data
        password = form.password.data
        protocol = form.protocol.data
        server = form.server.data
        port = form.port.data
        unique_filename = f'ftp_details_{uuid.uuid4().hex}.xlsx'
        output_file = os.path.join('static', unique_filename)

        try:
            if protocol == 'imap':
                emails = fetch_emails_imap(username, password, server, port)
            elif protocol == 'pop3':
                emails = fetch_emails_pop3(username, password, server, port)
            
            if not emails:
                flash("Failed to fetch emails.", 'danger')
                return jsonify({"status": "error", "message": "Failed to fetch emails."})
            else:
                details = extract_ftp_details_from_emails(emails)
                if details:
                    save_to_excel(details, output_file)
                    flash(f"Details saved to {output_file}", 'success')
                    return jsonify({"status": "success", "file": unique_filename})
                else:
                    flash("No details to save.", 'warning')
                    return jsonify({"status": "error", "message": "No details to save."})
        except RuntimeError as e:
            flash(str(e), 'danger')
            return jsonify({"status": "error", "message": str(e)})
    return render_template('index.html', form=form)

@app.route('/download/<filename>')
def download_file(filename):
    return send_from_directory('static', filename, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
