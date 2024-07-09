import imaplib
import poplib
import smtplib
import email
from email.policy import default
from bs4 import BeautifulSoup
import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import getpass

# IMAP Configuration
IMAP_SERVER = "outlook.office365.com"
IMAP_PORT = 993

# POP3 Configuration
POP_SERVER = "outlook.office365.com"
POP_PORT = 995

# SMTP Configuration
SMTP_SERVER = "smtp-mail.outlook.com"
SMTP_PORT = 587

def fetch_emails_imap(username, password):
    try:
        mail = imaplib.IMAP4_SSL(IMAP_SERVER, IMAP_PORT)
        mail.login(username, password)
        mail.select('inbox')
        status, data = mail.search(None, 'ALL')
        mail_ids = data[0].split()
        emails = []
        for i in tqdm(mail_ids, desc="Fetching emails via IMAP", unit="email"):
            status, data = mail.fetch(i, '(RFC822)')
            emails.append(data[0][1])
        mail.logout()
        return emails
    except imaplib.IMAP4.error as e:
        print(f"IMAP error: {e}")
        return []

def fetch_emails_pop3(username, password):
    try:
        mail = poplib.POP3_SSL(POP_SERVER, POP_PORT)
        mail.user(username)
        mail.pass_(password)
        num_messages = len(mail.list()[1])
        emails = []
        for i in tqdm(range(num_messages), desc="Fetching emails via POP3", unit="email"):
            response, lines, octets = mail.retr(i + 1)
            msg_data = b'\r\n'.join(lines)
            emails.append(msg_data)
        mail.quit()
        return emails
    except poplib.error_proto as e:
        print(f"POP3 error: {e}")
        return []

def parse_ftp_details_from_html(html_content):
    soup = BeautifulSoup(html_content, 'html.parser')
    account_info_table = soup.find('table', {'class': 'MsoNormalTable'})
    if not account_info_table:
        return None
    
    rows = account_info_table.find_all('tr')
    ftp_details = {}
    
    for row in rows:
        cells = row.find_all('td')
        if len(cells) == 3:
            label = cells[0].get_text(strip=True)
            key = cells[1].get_text(strip=True)
            value = cells[2].get_text(strip=True)
            if key and value:
                ftp_details[key] = value
    
    return ftp_details if ftp_details else None

def extract_ftp_details_from_emails(emails):
    all_details = []
    for raw_email in tqdm(emails, desc="Extracting FTP details", unit="email"):
        msg = email.message_from_bytes(raw_email, policy=default)
        for part in msg.walk():
            if part.get_content_type() == 'text/html':
                html_content = part.get_payload(decode=True)
                try:
                    html_content = html_content.decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        html_content = html_content.decode('latin-1')
                    except UnicodeDecodeError:
                        continue

                details = parse_ftp_details_from_html(html_content)
                if details:
                    details['Sender'] = msg['From']
                    details['Date'] = msg['Date']
                    all_details.append(details)
    return all_details

def save_to_excel(details_list, output_file):
    df = pd.DataFrame(details_list)
    df.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active

    # Adding dropdown for date column
    date_col = get_column_letter(df.columns.get_loc('Date') + 1)
    dv = DataValidation(type="list", formula1=f'"{",".join(df["Date"].unique())}"', showDropDown=True)
    ws.add_data_validation(dv)
    dv.add(f'{date_col}2:{date_col}{len(df) + 1}')

    wb.save(output_file)
    print(f"Details saved to {output_file}")

def automate_ftp_extraction(username, password, output_file):
    emails = fetch_emails_imap(username, password)
    if not emails:
        print("Failed to fetch emails via IMAP. Trying POP3...")
        emails = fetch_emails_pop3(username, password)
    if not emails:
        print("Failed to fetch emails via both IMAP and POP3.")
        return
    
    details = extract_ftp_details_from_emails(emails)
    if details:
        save_to_excel(details, output_file)
    else:
        print("No details to save.")

# Parameters
username = input("Enter your email username: ")
password = getpass.getpass("Enter your email password: ")
output_file = "ftp_details.xlsx"

# Execute the function
automate_ftp_extraction(username, password, output_file)

